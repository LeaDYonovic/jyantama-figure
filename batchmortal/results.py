import csv
import os
import re
import shutil

import openpyxl

CSV_COLUMNS = [
    "nickname",
    "source",
    "mode",
    "recordType",
    "uuid",
    "paipuUrl",
    "startTime",
    "endTime",
    "placement",
    "finalScore",
    "ptDelta",
    "playerLevel",
    "playerLevelScore",
    "resultUrl",
    "localPaipuPath",
    "modelTag",
    "rating",
    "errorMessage",
    "aiConsistencyRate",
    "aiConsistencyNumerator",
    "aiConsistencyDenominator",
    "temperature",
    "gameLength",
    "playerId",
    "reviewDuration",
    "screenshotPath",
    "timestamp",
    "badMoveRate5",
    "badMoveCount5",
    "badMoveRate10",
    "badMoveCount10",
    "badMoveDenominator",
]


def _backup_path(filepath: str) -> str:
    root, extension = os.path.splitext(filepath)
    return f"{root}.backup{extension}"


def _create_rolling_backup(filepath: str):
    if os.path.exists(filepath):
        shutil.copy2(filepath, _backup_path(filepath))


def _save_workbook_atomic(workbook, filepath: str):
    temporary_path = filepath + ".tmp"
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, filepath)
    finally:
        if os.path.exists(temporary_path):
            try:
                os.remove(temporary_path)
            except OSError:
                pass


def parse_metadata(metadata: dict) -> dict:
    """
    Parse the metadata dict returned by browser.py into typed fields.
    """

    def get(keys: list[str]) -> str:
        for raw_key, value in metadata.items():
            key = str(raw_key)
            lower_key = key.lower()
            for candidate in keys:
                if candidate.lower() in lower_key:
                    return value
        return ""

    ai_consistency = get(["\u4e00\u81f4\u7387", "Match Rate", "Matches/total"])
    numerator, denominator, rate = "", "", ""

    # Matches strings like "195/271 = 71.956%".
    match = re.search(r"(\d+)\s*/\s*(\d+)\s*=\s*([\d.]+)%", ai_consistency)
    if match:
        numerator = match.group(1)
        denominator = match.group(2)
        rate = match.group(3) + "%"

    return {
        "modelTag": get(["model tag"]),
        "rating": get(["rating"]),
        "aiConsistencyRate": rate,
        "aiConsistencyNumerator": numerator,
        "aiConsistencyDenominator": denominator,
        "temperature": get(["temperature", "\u6e29\u5ea6"]),
        "gameLength": get(["\u5bf9\u5c40\u957f\u5ea6", "length"]),
        "playerId": get(["\u73a9\u5bb6 ID", "player"]),
        "reviewDuration": get(["\u5ba1\u67e5\u7528\u65f6", "Duration"]),
    }


def backfill_result_metadata(
    filepath: str,
    output_format: str,
    metadata_by_uuid: dict[str, dict],
) -> int:
    """Fill blank imported-game fields without replacing completed analysis data."""
    if not metadata_by_uuid or not os.path.exists(filepath):
        return 0

    allowed_columns = {
        "mode",
        "recordType",
        "paipuUrl",
        "startTime",
        "endTime",
        "placement",
        "finalScore",
        "ptDelta",
        "playerLevel",
        "playerLevelScore",
    }
    updated_rows = 0

    def should_fill(column: str, current_value, new_value) -> bool:
        if new_value in (None, ""):
            return False
        current = str(current_value or "").strip()
        if not current:
            return True
        if column == "recordType" and current.lower() == "unknown":
            return str(new_value).strip().lower() != "unknown"
        if column == "mode" and current.lower() == "direct":
            return True
        return False

    if output_format == "csv":
        with open(filepath, "r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
        for row in rows:
            updates = metadata_by_uuid.get(str(row.get("uuid") or "").strip())
            if not updates:
                continue
            changed = False
            for column in allowed_columns:
                value = updates.get(column, "")
                if not should_fill(column, row.get(column), value):
                    continue
                row[column] = value
                changed = True
            updated_rows += int(changed)
        if updated_rows:
            _create_rolling_backup(filepath)
            with open(filepath, "w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)
        return updated_rows

    if output_format != "xlsx":
        raise ValueError(f"Unsupported output format: {output_format}")

    workbook = openpyxl.load_workbook(filepath)
    try:
        worksheet = workbook.active
        headers = {
            str(cell.value): cell.column
            for cell in worksheet[1]
            if cell.value is not None and str(cell.value)
        }
        for column in CSV_COLUMNS:
            if column not in headers:
                column_index = worksheet.max_column + 1
                worksheet.cell(row=1, column=column_index, value=column)
                headers[column] = column_index

        uuid_column = headers.get("uuid")
        if uuid_column is None:
            return 0
        for row_index in range(2, worksheet.max_row + 1):
            uuid = str(worksheet.cell(row=row_index, column=uuid_column).value or "").strip()
            updates = metadata_by_uuid.get(uuid)
            if not updates:
                continue
            changed = False
            for column in allowed_columns:
                value = updates.get(column, "")
                cell = worksheet.cell(row=row_index, column=headers[column])
                if not should_fill(column, cell.value, value):
                    continue
                cell.value = value
                changed = True
            updated_rows += int(changed)
        if updated_rows:
            _create_rolling_backup(filepath)
            _save_workbook_atomic(workbook, filepath)
    finally:
        workbook.close()
    return updated_rows


class ResultWriter:
    """
    Keep the output file open and flush in batches to avoid O(n^2) XLSX writes.
    """

    def __init__(self, filepath: str, output_format: str = "csv", flush_every: int = 1):
        self.filepath = filepath
        self.output_format = output_format
        self.flush_every = max(1, flush_every)
        self.is_new = not os.path.exists(filepath)
        self._pending_rows = 0
        self._file = None
        self._csv_writer = None
        self._workbook = None
        self._worksheet = None
        self._headers = []
        self._column_to_index = {}

        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        if not self.is_new:
            _create_rolling_backup(filepath)

        if output_format == "csv":
            if not self.is_new:
                self._migrate_csv_headers()
            self._file = open(filepath, mode="a", newline="", encoding="utf-8")
            self._csv_writer = csv.writer(self._file)
            if self.is_new:
                self._csv_writer.writerow(CSV_COLUMNS)
                self._file.flush()
        elif output_format == "xlsx":
            if self.is_new:
                self._workbook = openpyxl.Workbook()
                self._worksheet = self._workbook.active
                self._worksheet.append(CSV_COLUMNS)
                self._uuid_to_row = {}
                self._set_xlsx_headers(CSV_COLUMNS)
            else:
                self._workbook = openpyxl.load_workbook(filepath)
                self._worksheet = self._workbook.active
                self._uuid_to_row = {}
                headers = []
                for idx, row in enumerate(self._worksheet.iter_rows(values_only=True), start=1):
                    if idx == 1:
                        headers = [str(c) if c else "" for c in row]
                    else:
                        if "uuid" in headers:
                            uuid_idx = headers.index("uuid")
                            if len(row) > uuid_idx and row[uuid_idx]:
                                self._uuid_to_row[str(row[uuid_idx]).strip()] = idx
                if self._ensure_xlsx_headers(headers):
                    self._pending_rows += 1
                self._set_xlsx_headers(headers)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")

    def _migrate_csv_headers(self):
        try:
            with open(self.filepath, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames == CSV_COLUMNS:
                    return
                rows = list(reader)

            with open(self.filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(CSV_COLUMNS)
                for row in rows:
                    writer.writerow([row.get(column, "") for column in CSV_COLUMNS])
        except Exception as exc:
            raise RuntimeError(f"Failed to migrate CSV headers for {self.filepath}: {exc}") from exc

    def _ensure_xlsx_headers(self, headers: list[str]) -> bool:
        existing = set(headers)
        changed = False
        next_col = len(headers) + 1
        for column in CSV_COLUMNS:
            if column in existing:
                continue
            self._worksheet.cell(row=1, column=next_col, value=column)
            existing.add(column)
            headers.append(column)
            next_col += 1
            changed = True
        return changed

    def _set_xlsx_headers(self, headers: list[str]):
        self._headers = headers
        self._column_to_index = {
            column: idx
            for idx, column in enumerate(headers, start=1)
            if column
        }

    def _write_xlsx_row(self, row_idx: int, row: dict):
        for column in CSV_COLUMNS:
            col_idx = self._column_to_index[column]
            self._worksheet.cell(row=row_idx, column=col_idx, value=row.get(column, ""))

    def write_row(self, row: dict):
        safe_row = [row.get(column, "") for column in CSV_COLUMNS]

        if self.output_format == "csv":
            self._csv_writer.writerow(safe_row)
            self._pending_rows += 1
            if self._pending_rows >= self.flush_every:
                self.flush()
            return

        uuid_val = str(row.get("uuid", "")).strip()
        if uuid_val and hasattr(self, "_uuid_to_row") and uuid_val in self._uuid_to_row:
            row_idx = self._uuid_to_row[uuid_val]
            self._write_xlsx_row(row_idx, row)
        else:
            row_idx = self._worksheet.max_row + 1
            self._write_xlsx_row(row_idx, row)
            if uuid_val and hasattr(self, "_uuid_to_row"):
                self._uuid_to_row[uuid_val] = row_idx

        self._pending_rows += 1
        if self._pending_rows >= self.flush_every:
            self.flush()

    def flush(self):
        if self._pending_rows == 0:
            return

        if self.output_format == "csv":
            self._file.flush()
            os.fsync(self._file.fileno())
        else:
            _save_workbook_atomic(self._workbook, self.filepath)

        self._pending_rows = 0

    def close(self):
        try:
            self.flush()
        finally:
            if self._workbook is not None:
                self._workbook.close()
                self._workbook = None
                self._worksheet = None
            if self._file is not None:
                self._file.close()
                self._file = None
                self._csv_writer = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()


def append_row(filepath: str, row: dict, output_format: str = "csv"):
    """
    Backward-compatible one-shot append helper.
    """
    with ResultWriter(filepath, output_format=output_format, flush_every=1) as writer:
        writer.write_row(row)


def read_result_rows(filepath: str, output_format: str = "xlsx") -> list[dict]:
    if not os.path.exists(filepath):
        return []

    if output_format == "csv":
        with open(filepath, "r", encoding="utf-8") as f:
            return list(csv.DictReader(f))

    if output_format == "xlsx":
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = iter(ws.rows)
            first_row = next(rows, None)
            if first_row is None:
                return []

            headers = [str(cell.value) if cell.value is not None else "" for cell in first_row]
            result_rows = []
            for row in rows:
                record = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        record[headers[idx]] = cell.value if cell.value is not None else ""
                result_rows.append(record)
            return result_rows
        finally:
            wb.close()

    raise ValueError(f"Unsupported output format: {output_format}")


def get_processed_uuids(filepath: str, output_format: str = "xlsx") -> set[str]:
    """
    Return UUIDs with a non-empty, non-ERROR Rating as completed analyses.

    This includes INVALID rows: those inputs were permanently rejected by the
    review service and should not consume time again on later runs.
    """
    processed = set()
    if not os.path.exists(filepath):
        return processed

    try:
        if output_format == "csv":
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if "uuid" in row and row["uuid"]:
                        rating = str(row.get("rating") or "").strip()
                        if rating and rating.upper() != "ERROR":
                            processed.add(row["uuid"])
        elif output_format == "xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            try:
                ws = wb.active
                rows = iter(ws.rows)
                first_row = next(rows, None)
                if first_row is not None:
                    headers = [cell.value for cell in first_row]
                    if "uuid" in headers:
                        uuid_idx = headers.index("uuid")
                        rating_idx = headers.index("rating") if "rating" in headers else -1
                        for row in rows:
                            if len(row) <= uuid_idx or not row[uuid_idx].value:
                                continue
                            rating_val = (
                                str(row[rating_idx].value).strip()
                                if rating_idx >= 0
                                and len(row) > rating_idx
                                and row[rating_idx].value is not None
                                else ""
                            )
                            if rating_val and rating_val.upper() != "ERROR":
                                processed.add(str(row[uuid_idx].value).strip())
            finally:
                wb.close()
    except Exception as e:
        print(f"Failed to read processed UUIDs from {filepath}: {e}")
        
    return processed
